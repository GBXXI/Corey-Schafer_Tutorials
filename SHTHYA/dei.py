#------------------DEI-------------------

#------------------DEI 1-----------------

import pandas as pd
from openpyxl import *
import numpy as np
import matplotlib.pyplot as plt
plt.close('all')


#dhmioyrgoume enan pianaka me hmeromhnies wra pou mas afora gia dei
m=int(input("poses meres: "))
#date=input("prwth mhnos: ")
#hmeromhnies = pd.date_range(date, periods=m, freq='D')
#hmeromhnies.to_series(keep_tz=True,name="hmeromhnies2")

sr = pd.Series(['2021-10-01 07:00:00', '2021-10-02 07:00:00', 
                '2021-10-03 07:00:00', '2021-10-04 07:00:00', 
                '2021-10-05 07:00:00', '2021-10-06 07:00:00', 
                '2021-10-07 07:00:00', '2021-10-08 07:00:00', 
                '2021-10-09 07:00:00', '2021-10-10 07:00:00',
                '2021-10-11 07:00:00', '2021-10-12 07:00:00', 
                '2021-10-13 07:00:00', '2021-10-14 07:00:00', 
                '2021-10-15 07:00:00', '2021-10-16 07:00:00',
                '2021-10-17 07:00:00', '2021-10-18 07:00:00',
                '2021-10-19 07:00:00', '2021-10-20 07:00:00', 
                '2021-10-21 07:00:00', '2021-10-22 07:00:00', 
                '2021-10-23 07:00:00', '2021-10-24 07:00:00',
                '2021-10-25 07:00:00', '2021-10-26 07:00:00', 
                '2021-10-27 07:00:00', '2021-10-28 07:00:00', 
                '2021-10-29 07:00:00', '2021-10-30 07:00:00',
                '2021-10-31 07:00:00',"0"])


dei1=pd.read_excel("dei1.xlsx")
dimensions=dei1.shape
wbDEI=load_workbook('dei1.xlsx')
wsDEI=wbDEI["Sheet1"]
wb=load_workbook('a.xlsx')
ws=wb['b']

#to q einai o deikths toy stoixeiou tou arxikopoihmenou pinaka me hmeromhnies/wra
q=0
print("dei1")
for row in range (1,3092):
       for col in range(1,85):
#se poio cell briskomai sto excel ths deh basei twn apo panw for
             wcellDEI=wsDEI.cell(row,col)
#an to keli sto excel ths deh exei to idio onoma me to q stoixeio ston arxikopoihmeno pinaka sr me tis hmeromhnies/wra
#kai an eimaste akoma stis 30 meres      
             if wcellDEI.value==sr.loc[q] and q<=31:
                 for s in range (2,6):
                    wcellDEI=wsDEI.cell(row,col+s)
                    tempcell=wcellDEI.value
                    print(tempcell)
                    if s==2:
                        wcell=ws.cell(q+3,56)
                        wcell.value=tempcell
                    elif s==3:
                        wcell=ws.cell(q+3,62)
                        wcell.value=tempcell
                    elif s==4:
                        wcell=ws.cell(q+3,68)
                        wcell.value=tempcell
                    elif s==5:
                        wcell=ws.cell(q+3,74)
                        wcell.value=tempcell
                 q=q+1
wb.save('a.xlsx')                 
  

#------------------DEI 2-----------------
       
dei2=pd.read_excel("dei2.xlsx")
dimensions=dei1.shape
wbDEI=load_workbook('dei2.xlsx')
wsDEI=wbDEI["Sheet1"]
wb=load_workbook('a.xlsx')
ws=wb['b']

#to q einai to row (hmera) sto excel anaforas
q=0
print("dei2")
for row in range (1,3092):
       for col in range(1,85):
             wcellDEI=wsDEI.cell(row,col)
             if wcellDEI.value==sr.loc[q] and q<=31:
                 for s in range (2,6):
                    wcellDEI=wsDEI.cell(row,col+s)
                    tempcell=wcellDEI.value
                    print(tempcell)
                    if s==2:
                        wcell=ws.cell(q+3,58)
                        wcell.value=tempcell
                    elif s==3:
                        wcell=ws.cell(q+3,64)
                        wcell.value=tempcell
                    elif s==4:
                        wcell=ws.cell(q+3,70)
                        wcell.value=tempcell
                    elif s==5:
                        wcell=ws.cell(q+3,76)
                        wcell.value=tempcell
                 q=q+1
wb.save('a.xlsx')                 