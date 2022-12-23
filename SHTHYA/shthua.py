# https://altoconvertpdftoexcel.com/ kai smallpdf (prosoxh poiasthlh (h g na einai nomizw))
import pandas as pd
from openpyxl import *
import numpy as np
import matplotlib.pyplot as plt
plt.close('all')



#-----------------SHTUA------------------

#orizw excel kai sheet
wb=load_workbook('a.xlsx')
ws=wb['b']

#dhmiourgw list me ta onomata twn excel pou tha kanw read
meres=["1.xlsx","2.xlsx","3.xlsx","4.xlsx","5.xlsx","6.xlsx","7.xlsx",
       "8.xlsx","9.xlsx","10.xlsx","11.xlsx","12.xlsx","13.xlsx",
       "14.xlsx","15.xlsx","16.xlsx","17.xlsx","18.xlsx","19.xlsx",
       "20.xlsx","21.xlsx","22.xlsx","23.xlsx","24.xlsx","25.xlsx",
       "26.xlsx","27.xlsx","28.xlsx","29.xlsx","30.xlsx","31.xlsx"] 

#prepei na kserw poses meres exei o mhnas kai poia mera den exw anafora na dhmiourghsw mhdenikh
m=int(input("poses meres: "))
k=0
l=0
k=int(input("poia oxi 1: "))
l=int(input("poia oxi 2: "))


#loupa gia na kanw read ola ta excel (ksekinaei apo 0 ftanei mexri m-1)
for n in range (0,m):
#synthiki an briskomaste sth lοupa pou den expume anafora gia na dhmiourghsoume mhdenikh
#gia perissoteres meres tis prosthetw me or kai diabazw perissoteres apo panw
       if n==k-1 or n==l-1:
           print("den exw report")
           df1 = pd.Series(0, index =[10, 11, 14, 15, 22, 23, 33, 37, 49]) 
#diaforetika synexizw kanonika ta read
       else:
           print(meres[n])
#reaf to excel
           dfa=pd.read_excel(meres[n])
#eksagw olh th sthlh pou me endiaferei (me tis times)
           dfb=dfa.iloc[:,5]
#eksagw ta stoixeia pou me endiaferoyn apo olh th sthlh poy phra
           df1=dfb[[x for x in [10, 11, 14, 15, 22, 23, 33, 37, 49]]]
#epeidh kanw synenwsh pinakwn prepei na ksekinhsw apo kapoion. etsi an einai prwth fora
#o pinakas mou einai ta stoixeia ths sthlhs pou phra apo to 1o excel
       if n==0:
               pinakas = df1
#diaforetika enwnw kathe proigoumeno pinakas me ta nea stoixeia sthlhs pou exw eksagei
       else:
               pinakas = pd.concat([pinakas,df1], axis=1)
               
#onomazw ta headers twn sthlwn gt bgainoun unnamed kanonika kai prepei na tous dwsw
#onomata gia na mporw na exw prosbash
pinakas.columns = [1,2,3,4,5,6,7,8,9,10,11,12,13,14,15,16,17,18,19,20,21,22,23,24,25,26,27,
                   28,29,30]
#megethi=["h1","h2","p1","p2","q1","q2","mj","nm3"]
#pinakas.index=[megethi]
#metrhths pou auksanetai se kathe loupa gia na odhgoumai stis sthles tou "pinaka" opou
#kathe sthlh einai mia mera         
y=1
print(pinakas)
#to x apeikonizei th mera (dld th seira) sto excel anaforas
for x in range(4,m+4):
#for x in megethi:
#pernaw tis wres leitourgias twn gennhtriwn
#epilegw to keli sto opoio tha eisagw timh
      wcell=ws.cell(x,2)
      #eisagw timh apo ton "pinaka" 
      wcell.value=pinakas.loc[10,y]
      wcell=ws.cell(x,3)
      wcell.value=pinakas.loc[11,y]
      #pernaw tis mwh
      wcell=ws.cell(x,6)
      wcell.value=pinakas.loc[14,y]/1000
      wcell=ws.cell(x,7)
      wcell.value=pinakas.loc[15,y]/1000
      #pernaw tis mvarh
      wcell=ws.cell(x,9)
      wcell.value=pinakas.loc[22,y]/1000
      wcell=ws.cell(x,10)
      wcell.value=pinakas.loc[23,y]/1000
      #pernaw nm3 kai mj aeriou
      wcell=ws.cell(x,23)
      wcell.value=pinakas.loc[37,y]
      wcell=ws.cell(x,24)
      wcell.value=pinakas.loc[33,y]
      #pernaw thermal energy
      wcell=ws.cell(x,36)
      wcell.value=pinakas.loc[49,y]
      y=y+1

#swse to excel anaforas me tis times pou edwsa
wb.save('a.xlsx')

###pinakas.plot();
#ypologismoi=pd.Dataframe()
#ypologismoi=pinakas
#del ypologismoi[k]
#maxValues = ypologismoi.max(axis = 1) 
#minValues = ypologismoi.min(axis=1)
#minValues=pd.concat([minValues,maxValues], axis=1)
#minValues.columns = ["min","max"]
#minValues.index=["h1","h2","p1","p2","q1","q2","mj","nm3","scv"]
#print(minValues)


